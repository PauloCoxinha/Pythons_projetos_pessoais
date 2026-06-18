import openpyxl
from openpyxl.styles import Font, PatternFill

# 1. Criar um novo arquivo ou carregar um existente
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = "Meus Animes"

# 2. Criar o cabeçalho
headers = ["Anime", "Gênero", "Prioridade"]
sheet.append(headers)

# 3. Adicionar dados (exemplo baseado na sua planilha de não assistidos)
#
dados = [
    ["Death Note #151", "Fantasia", "Alta"],
    ["Vinland Saga #152", "Ação", "Média"],
    ["Kuroko no Basket #153", "Fantasia", "Baixa"]
]

for linha in dados:
    sheet.append(linha)

# 4. Estilizar o cabeçalho (Negrito e Cor de Fundo)
header_fill = PatternFill(start_color="FF9900", end_color="FF9900", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")

for cell in sheet[1]:
    cell.fill = header_fill
    cell.font = header_font

# 5. Salvar o arquivo
wb.save("Planilha_Formatada_Animes.xlsx")
print("Planilha .xlsx criada com sucesso!")